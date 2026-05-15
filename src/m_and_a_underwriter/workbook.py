"""openpyxl workbook export. Sources & Uses, projection, returns, sensitivity."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .schema import Inputs, ModelOutput

BOLD = Font(bold=True)
HEADER_FILL = PatternFill("solid", fgColor="0A0A0A")
HEADER_FONT = Font(bold=True, color="FFFFFF")
RUST_FILL = PatternFill("solid", fgColor="FF4D2E")


def _autosize(ws):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col_letter].width = min(max(width + 2, 12), 40)


def export(inp: Inputs, model: ModelOutput, path: str | Path) -> Path:
    path = Path(path)
    wb = Workbook()

    # Cover
    ws = wb.active
    ws.title = "Cover"
    ws["A1"] = f"{inp.target.name} \u2014 LBO Model"
    ws["A1"].font = Font(bold=True, size=20)
    ws["A3"] = "Industry"
    ws["B3"] = inp.target.industry
    ws["A4"] = "Entry multiple"
    ws["B4"] = inp.deal.entry_multiple
    ws["A5"] = "Exit multiple"
    ws["B5"] = inp.deal.exit_multiple
    ws["A6"] = "Hold years"
    ws["B6"] = inp.deal.hold_years
    ws["A8"] = "MOIC"
    ws["B8"] = round(model.returns.moic, 2)
    ws["A9"] = "IRR"
    ws["B9"] = f"{model.returns.irr*100:.1f}%"
    for cell in ("A3", "A4", "A5", "A6", "A8", "A9"):
        ws[cell].font = BOLD
    _autosize(ws)

    # Sources & Uses
    su_ws = wb.create_sheet("Sources_Uses")
    su = model.sources_uses
    rows = [
        ("USES", ""),
        ("Purchase EV", su.purchase_ev),
        ("Refi existing debt", su.refinanced_debt),
        ("Transaction fees", su.transaction_fees),
        ("Financing fees", su.financing_fees),
        ("Minimum cash", su.minimum_cash),
        ("Total Uses", su.total_uses),
        ("", ""),
        ("SOURCES", ""),
        ("Total Debt", su.total_debt),
        ("Sponsor Equity", su.sponsor_equity),
        ("Total Sources", su.total_sources),
    ]
    for i, (k, v) in enumerate(rows, 1):
        su_ws.cell(row=i, column=1, value=k)
        su_ws.cell(row=i, column=2, value=v)
        if k in ("USES", "SOURCES"):
            su_ws.cell(row=i, column=1).fill = HEADER_FILL
            su_ws.cell(row=i, column=1).font = HEADER_FONT
        if k in ("Total Uses", "Total Sources"):
            su_ws.cell(row=i, column=1).font = BOLD
            su_ws.cell(row=i, column=2).font = BOLD
    _autosize(su_ws)

    # Projection
    proj_ws = wb.create_sheet("Projection")
    headers = [
        "Year", "Revenue", "EBITDA", "D&A", "EBIT", "Interest", "Pretax", "Taxes", "Net Income",
        "Capex", "Δ NWC", "CFO pre-debt", "Mandatory Amort", "Cash Sweep", "Debt Paydown",
        "Beg Debt", "End Debt", "Beg Cash", "End Cash",
    ]
    for col, h in enumerate(headers, 1):
        c = proj_ws.cell(row=1, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
    for r, p in enumerate(model.projections, 2):
        vals = [p.year, p.revenue, p.ebitda, p.d_and_a, p.ebit, p.interest, p.pretax, p.taxes,
                p.net_income, p.capex, p.delta_nwc, p.cash_flow_pre_debt, p.mandatory_amort,
                p.cash_sweep, p.debt_paydown, p.beg_debt, p.end_debt, p.beg_cash, p.end_cash]
        for col, v in enumerate(vals, 1):
            proj_ws.cell(row=r, column=col, value=round(v, 2) if isinstance(v, float) else v)
    _autosize(proj_ws)

    # Returns
    ret_ws = wb.create_sheet("Returns")
    r = model.returns
    rows = [
        ("Exit EBITDA", r.exit_ebitda),
        ("Exit multiple", inp.deal.exit_multiple),
        ("Exit EV", r.exit_ev),
        ("Exit equity", r.exit_equity),
        ("MOIC", round(r.moic, 2)),
        ("IRR", f"{r.irr*100:.1f}%"),
    ]
    for i, (k, v) in enumerate(rows, 1):
        ret_ws.cell(row=i, column=1, value=k).font = BOLD
        ret_ws.cell(row=i, column=2, value=v)
        if k == "IRR":
            ret_ws.cell(row=i, column=2).fill = RUST_FILL
            ret_ws.cell(row=i, column=2).font = BOLD
    _autosize(ret_ws)

    # Sensitivity grid
    sens_ws = wb.create_sheet("Sensitivity")
    sens_ws.cell(row=1, column=1, value="Entry \\ Exit").font = BOLD
    rows = list(model.sensitivity.items())
    exit_keys = list(rows[0][1].keys()) if rows else []
    for col, k in enumerate(exit_keys, 2):
        c = sens_ws.cell(row=1, column=col, value=k)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")
    for r_idx, (entry_k, row) in enumerate(rows, 2):
        sens_ws.cell(row=r_idx, column=1, value=entry_k).font = BOLD
        for col, k in enumerate(exit_keys, 2):
            v = row.get(k, 0)
            cell = sens_ws.cell(row=r_idx, column=col, value=f"{v*100:.1f}%")
            cell.alignment = Alignment(horizontal="center")
    _autosize(sens_ws)

    wb.save(path)
    return path
